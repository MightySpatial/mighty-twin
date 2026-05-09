"""Admin upload parser — Phase F.

Server-side ingestion of CSV / GeoJSON / XLSX / Gaussian-splat into a
DataSource row. Files land in /tmp/twin-uploads/ (good enough for local
dev; Railway deploy switches to S3-compatible object storage in Phase
J).

Pipeline:
  1. multipart/form-data POST to /api/upload  (admin only)
  2. detect type by extension (csv | geojson | xlsx | splat)
  3. parse → infer attribute schema (columns + types) for tabular formats;
     for splats, sniff the format flavour (PlayCanvas / Pix4D / generic)
  4. persist DataSource row referencing the on-disk file
  5. return the new DataSource so the frontend can wire a Layer to it
"""

from __future__ import annotations

import csv
import io
import json
import os
import struct
import uuid
from pathlib import Path
from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status

from mighty_models import DataSource

from .auth import get_current_user, require_admin
from .db import get_db

UPLOAD_DIR = Path(os.environ.get("TWIN_UPLOAD_DIR", "/tmp/twin-uploads"))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

router = APIRouter(tags=["upload"])

MAX_BYTES = 50 * 1024 * 1024  # 50 MB cap for tabular uploads
# Gaussian splat files are an order of magnitude bigger — half a million
# splats (a small building scan) is already ~50MB raw, full PlayCanvas /
# Pix4D outputs commonly land in the 200–500MB range.
SPLAT_MAX_BYTES = 600 * 1024 * 1024
SPLAT_EXTS = {"splat", "ply", "spz", "ksplat", "compressed.ply"}


def _detect_ext(fname: str) -> str:
    """Lower-case extension. Handles ``.compressed.ply`` (PlayCanvas) as a
    distinct value rather than collapsing it to ``ply``."""
    lower = fname.lower()
    if lower.endswith(".compressed.ply"):
        return "compressed.ply"
    return lower.rsplit(".", 1)[-1] if "." in lower else ""


@router.post("/api/upload", status_code=201)
async def upload(
    file: Annotated[UploadFile, File(...)],
    name: Annotated[str | None, Form()] = None,
    _admin = Depends(lambda u=Depends(get_current_user): require_admin(u)),
    db = Depends(get_db),
) -> dict[str, Any]:
    raw = await file.read()
    fname = file.filename or "upload"
    ext = _detect_ext(fname)
    is_splat = ext in SPLAT_EXTS

    cap = SPLAT_MAX_BYTES if is_splat else MAX_BYTES
    if len(raw) > cap:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File exceeds {cap // (1024*1024)}MB cap",
        )

    parsed: dict[str, Any]
    type_: str
    if is_splat:
        parsed = _sniff_splat(raw, ext, fname)
        type_ = "splat"
    elif ext in {"geojson", "json"}:
        parsed = _parse_geojson(raw)
        type_ = "geojson"
    elif ext == "csv":
        parsed = _parse_csv(raw)
        type_ = "csv"
    elif ext in {"xlsx", "xlsm"}:
        parsed = _parse_xlsx(raw)
        type_ = "xlsx"
    else:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=(
                f"Unsupported extension {ext!r}; expected csv/geojson/xlsx "
                "or a Gaussian splat (.ply, .compressed.ply, .splat, .spz, .ksplat)"
            ),
        )

    uid = uuid.uuid4()
    out_path = UPLOAD_DIR / f"{uid}-{fname}"
    out_path.write_bytes(raw)

    attributes: dict[str, Any]
    if is_splat:
        # No tabular schema for binary splats — store provenance + format
        # so downstream rendering can pick the right loader.
        attributes = {
            "kind": "splat",
            "format": parsed.get("format"),
            "splat_format": parsed.get("splat_format"),
            "splat_count": parsed.get("splat_count"),
            "origin_hint": parsed.get("origin_hint"),
            "filename": fname,
        }
    else:
        attributes = {
            "schema": parsed.get("schema", []),
            "feature_count": parsed.get("count", 0),
            "preview": parsed.get("preview", []),
        }

    ds = DataSource(
        id=uid,
        name=name or fname,
        type=type_,
        url=str(out_path),
        size_bytes=len(raw),
        attributes=attributes,
    )
    db.add(ds)
    db.commit()
    db.refresh(ds)

    return {
        "id": str(ds.id),
        "name": ds.name,
        "type": ds.type,
        "url": ds.url,
        "size_bytes": ds.size_bytes,
        "attributes": ds.attributes,
    }


# ── Parsers ─────────────────────────────────────────────────────────────


def _parse_geojson(raw: bytes) -> dict[str, Any]:
    try:
        doc = json.loads(raw)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid GeoJSON: {e}")
    features = doc.get("features", []) if isinstance(doc, dict) else []
    schema_keys: dict[str, str] = {}
    for f in features[:100]:
        for k, v in (f.get("properties") or {}).items():
            schema_keys.setdefault(k, _typeof(v))
    schema = [{"name": k, "type": t} for k, t in schema_keys.items()]
    return {
        "schema": schema,
        "count": len(features),
        "preview": features[:5],
    }


def _parse_csv(raw: bytes) -> dict[str, Any]:
    text = raw.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    schema = [{"name": h, "type": "text"} for h in (reader.fieldnames or [])]
    for col in schema:
        sampled = [r.get(col["name"], "") for r in rows[:50]]
        col["type"] = _column_type(sampled)
    return {"schema": schema, "count": len(rows), "preview": rows[:5]}


def _parse_xlsx(raw: bytes) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return {"schema": [], "count": 0, "preview": []}
    rows_iter = ws.iter_rows(values_only=True)
    headers_row = next(rows_iter, ())
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(headers_row)]
    rows: list[dict[str, Any]] = []
    for r in rows_iter:
        rows.append({h: v for h, v in zip(headers, r)})
    schema = [
        {"name": h, "type": _column_type([r.get(h) for r in rows[:50]])}
        for h in headers
    ]
    return {"schema": schema, "count": len(rows), "preview": rows[:5]}


def _sniff_splat(raw: bytes, ext: str, fname: str) -> dict[str, Any]:
    """Identify a Gaussian splat file by extension + magic header.

    Recognised flavours:

    * ``ply`` — standard PLY ASCII or binary. Used by Pix4D's RealityCapture
      / Pix4DSplat exports and by raw outputs from gsplat trainers. Header
      starts with ``ply\\n``.
    * ``compressed.ply`` — PlayCanvas SuperSplat compressed PLY (their
      single-file deliverable from playcanvas/supersplat). Same PLY framing
      with custom property names like ``packed_position`` / ``packed_rot``.
    * ``splat`` — antimatter15/Niantic spec, 32 bytes per splat (xyz f32 +
      scale f32x3 + rgba u8x4 + rot u8x4). Pure binary, no header.
    * ``spz`` — Niantic Scaniverse compressed splat (gzipped binary with
      a 16-byte header magic ``\\x1f\\x8b`` after content).
    * ``ksplat`` — Mark Kellogg's gaussian-splats-3d packed format (binary
      with ASCII magic ``KSPLAT`` at offset 0).

    Heuristics here are conservative: when we can't tell, we record the
    ext as the format and leave splat_count null. The viewer-side loader
    is the source of truth for correctness — this just lights up the
    Atlas UI with provenance.
    """
    head = raw[:16]
    sample = raw[:4096]
    out: dict[str, Any] = {
        "format": ext,
        "splat_format": ext,
        "splat_count": None,
        "origin_hint": None,
    }

    if ext in {"ply", "compressed.ply"}:
        if not head.startswith(b"ply"):
            raise HTTPException(
                status_code=400,
                detail=f"{fname!r} doesn't have a PLY magic header",
            )
        # PLY headers are ASCII until the ``end_header`` line.
        try:
            # Decode only up to end_header — the body may be binary.
            text_head = sample.decode("ascii", errors="replace")
        except Exception:
            text_head = ""
        end = text_head.find("end_header")
        header_text = text_head[: end + len("end_header")] if end >= 0 else text_head
        # Splat-count is the first ``element vertex N`` line (or
        # ``element point N`` for some Pix4D variants).
        for line in header_text.splitlines():
            ll = line.strip()
            if ll.startswith("element vertex ") or ll.startswith("element point "):
                try:
                    out["splat_count"] = int(ll.rsplit(" ", 1)[1])
                except ValueError:
                    pass
                break
        # Origin hint — PlayCanvas SuperSplat compressed PLY stamps a
        # ``comment generator: SuperSplat`` style line; Pix4D stamps
        # ``comment Generated by Pix4D``. Use whichever we find first.
        h_lower = header_text.lower()
        if "supersplat" in h_lower or "playcanvas" in h_lower:
            out["origin_hint"] = "playcanvas"
        elif "pix4d" in h_lower:
            out["origin_hint"] = "pix4d"
        elif "scaniverse" in h_lower:
            out["origin_hint"] = "scaniverse"
        elif "polycam" in h_lower:
            out["origin_hint"] = "polycam"
        # Disambiguate compressed.ply by ``packed_position`` property name.
        if "packed_position" in h_lower:
            out["splat_format"] = "compressed.ply"
            out.setdefault("origin_hint", "playcanvas")

    elif ext == "splat":
        # Antimatter spec: 32 bytes per splat. Reject anything that's
        # plainly not a multiple — guards against accidental .splat from
        # other apps.
        if len(raw) % 32 != 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"{fname!r} is not a multiple of 32 bytes — not an "
                    "antimatter15-spec .splat file"
                ),
            )
        out["splat_count"] = len(raw) // 32
        out["origin_hint"] = "antimatter15"

    elif ext == "spz":
        # Scaniverse SPZ. Magic check is best-effort; the spec is gzip so
        # the second byte is 0x8b. Don't decompress here — the viewer
        # loader does that.
        if len(raw) < 4 or raw[0] != 0x1F or raw[1] != 0x8B:
            # Not strictly fatal; SPZ v2 may differ. Mark as unknown.
            out["origin_hint"] = "scaniverse?"
        else:
            out["origin_hint"] = "scaniverse"

    elif ext == "ksplat":
        if not head.startswith(b"KSPLAT"):
            raise HTTPException(
                status_code=400,
                detail=f"{fname!r} doesn't have a KSPLAT header",
            )
        # gaussian-splats-3d packs splat count at offset 6 as little-endian
        # uint32.
        try:
            (count,) = struct.unpack_from("<I", raw, 6)
            out["splat_count"] = count
        except struct.error:
            pass
        out["origin_hint"] = "kellogg"

    return out


def _typeof(v: Any) -> str:
    if isinstance(v, bool):
        return "boolean"
    if isinstance(v, int):
        return "integer"
    if isinstance(v, float):
        return "real"
    return "text"


def _column_type(sample: list[Any]) -> str:
    nonblank = [v for v in sample if v not in (None, "")]
    if not nonblank:
        return "text"
    try:
        for v in nonblank:
            int(str(v))
        return "integer"
    except (ValueError, TypeError):
        pass
    try:
        for v in nonblank:
            float(str(v))
        return "real"
    except (ValueError, TypeError):
        pass
    if all(str(v).lower() in {"true", "false", "0", "1", "yes", "no"} for v in nonblank):
        return "boolean"
    return "text"
